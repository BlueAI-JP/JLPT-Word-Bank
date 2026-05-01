from pathlib import Path
from typing import Optional
import openpyxl

BASE_DIR = Path(__file__).parent

LEVEL_CONFIG: dict[str, dict] = {
    "N4": {
        "xlsx": BASE_DIR / "WordBank" / "JLPTWordBank_N4_ALL.xlsx",
        "audio_dir": BASE_DIR / "JLPT_N4",
    },
    "N3": {
        "xlsx": BASE_DIR / "WordBank" / "JLPTWordBank_N3_ALL.xlsx",
        "audio_dir": BASE_DIR / "JLPT_N3",
    },
    "N2": {
        "xlsx": BASE_DIR / "WordBank" / "JLPTWordBank_N2_ALL.xlsx",
        "audio_dir": BASE_DIR / "JLPT_N2",
    },
    "N1": {
        "xlsx": BASE_DIR / "WordBank" / "JLPTWordBank_N1_ALL.xlsx",
        "audio_dir": BASE_DIR / "JLPT_N1",
    },
}


class WordDataLoader:
    def __init__(self) -> None:
        self._cache: dict[str, list[dict]] = {}

    def load_all(self) -> None:
        for level, cfg in LEVEL_CONFIG.items():
            if cfg["xlsx"].exists():
                self._cache[level] = self._load_xlsx(cfg["xlsx"])

    def _load_xlsx(self, path: Path) -> list[dict]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Normalize column names (N3 uses '單字的日文假名', N4 uses '日文假名')
        col_map = {
            "單字編號": "id",
            "日文單字": "word",
            "日文假名": "reading",
            "單字的日文假名": "reading",
            "詞性": "pos",
            "繁體中文翻譯": "zh",
            "日文例句": "example",
            "日文例句繁中翻譯": "example_zh",
        }

        words = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            entry: dict = {}
            for i, header in enumerate(headers):
                key = col_map.get(header, header)
                entry[key] = row[i]
            words.append(entry)

        wb.close()
        return words

    def get_available_levels(self) -> list[str]:
        return [lvl for lvl in LEVEL_CONFIG if lvl in self._cache]

    def get_all_levels(self) -> list[dict]:
        result = []
        for level in ["N4", "N3", "N2", "N1"]:
            available = level in self._cache
            result.append({
                "level": level,
                "available": available,
                "total": len(self._cache[level]) if available else 0,
            })
        return result

    def get_words(self, level: str) -> list[dict]:
        return self._cache.get(level, [])

    def get_word_by_id(self, level: str, word_id: int) -> Optional[dict]:
        for w in self._cache.get(level, []):
            if w["id"] == word_id:
                return w
        return None

    def get_audio_path(self, level: str, word_id: int) -> Optional[Path]:
        cfg = LEVEL_CONFIG.get(level)
        if cfg is None:
            return None
        path = cfg["audio_dir"] / f"{word_id:04d}.mp3"
        return path if path.exists() else None

    def level_has_audio(self, level: str) -> bool:
        cfg = LEVEL_CONFIG.get(level)
        if cfg is None:
            return False
        return cfg["audio_dir"].exists()
